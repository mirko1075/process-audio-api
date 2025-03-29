import json
from flask import Flask, after_this_request, request, jsonify, send_file
import os
import logging
import sys
from openpyxl import Workbook # type: ignore
from distutils.util import strtobool # type: ignore
from functools import wraps
from dotenv import load_dotenv
import io
import datetime
import time
from flask import g
import requests
from pdf_generator import PDFGenerator
from process_audio import RATE_PER_MINUTE, create_word_document, convert_to_wav, create_sentiment_details_df, create_sentiment_summary_df, delete_from_gcs, generate_multi_sheet_excel, get_audio_duration_from_form_file, get_usage_data, load_excel_file, log_audio_processing, parse_glossary_and_corrections, process_queries, transcribe_with_deepgram, transcript_with_whisper_large_files, transcribe_audio_openai, translate_text_google, translate_text_with_deepseek, translate_text_with_openai, upload_to_gcs # Ensure this uses the updated process_audio.py
from sentiment_analysis import run_sentiment_analysis
import assemblyai as aai

# Load environment variables
load_dotenv()

# Check for Google Cloud credentials at startup
GOOGLE_CREDENTIALS_PATH = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")

if not GOOGLE_CREDENTIALS_PATH:
    logging.error("ERROR: GOOGLE_APPLICATION_CREDENTIALS environment variable is not set.")
    sys.exit(1)  # Exit if credentials are not set

if not os.path.exists(GOOGLE_CREDENTIALS_PATH):
    logging.error(f"ERROR: Credentials file not found at {GOOGLE_CREDENTIALS_PATH}")
    sys.exit(1)  # Exit if file does not exist

logging.info(f"Google Cloud credentials found at {GOOGLE_CREDENTIALS_PATH}")

app = Flask(__name__)
# Configure logging
logging.basicConfig(level=logging.INFO)

API_KEY = os.getenv('API_KEY')
ASSEMBLYAI_API_KEY = os.getenv('ASSEMBLYAI_API_KEY')

aai.settings.api_key = ASSEMBLYAI_API_KEY

if not API_KEY:
    raise ValueError("API_KEY must be set in .env file")

def is_valid_json(json_string):
    try:
        json.loads(json_string)
        return True
    except json.JSONDecodeError:
        return False


def require_api_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = request.headers.get('x-api-key')
        if api_key and api_key == API_KEY:
            return f(*args, **kwargs)
        logging.error("Invalid or missing API key")
        return jsonify({"error": "Invalid or missing API key"}), 401
    return decorated

@app.before_request
def start_timer():
    g.start_time = time.time()

@app.after_request
def log_execution_time(response):
    if hasattr(g, 'start_time'):
        execution_time = time.time() - g.start_time
        print(f"Endpoint: {request.path} | Method: {request.method} | Time: {execution_time:.4f} sec")
        response.headers["X-Execution-Time"] = str(execution_time)  # Optional: Add to headers
    return response

@app.route("/", methods=["GET"])
def home():
    return jsonify({"message": "Hello, World!"})


@app.route('/transcribe_and_translate', methods=['POST'])
@require_api_key
def transcribe_and_translate():
    """Handles audio transcription and translation"""
    try:
        logging.info("Received request to transcribe and translate audio")
        
        audio_file = request.files.get("audio")
        translate = strtobool(request.form.get("translate", 'false'))
        translation_model = request.form.get("translation_model", "google")
        #accept google or openai
        if translate and translation_model not in ["google", "openai"]:
            return jsonify({"error": "Invalid translation model"}), 400

        language = request.form.get("sourceLanguage", "en")
        target_language = request.form.get("targetLanguage", "en")

        if not audio_file:
            return jsonify({"error": "No file uploaded"}), 400

        # Process and transcribe audio
        transcription_response = transcribe_with_deepgram(audio_file, language)
        logging.info(f"TRANSCRIPTION RESPONSE: {transcription_response}")
        formatted_transcript_array = transcription_response["formatted_transcript_array"]
        transcript = transcription_response["transcript"]
        translated_text = None
        if translate:
            # Send transcript for translation
            translated_text = translate_text_with_openai(transcript, language, target_language)

        # Return JSON response
        return jsonify({
            "formatted_transcript_array": formatted_transcript_array,
            "transcript": transcript,
            "translated_text": translated_text,
        })

    except Exception as e:
        logging.error(f"Error in transcribe_and_translate: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/get-audio-duration", methods=["POST"])
@require_api_key
def get_audio_duration_endpoint():
    """
    Endpoint to upload an audio file and return its duration.
    """
    audio_file = request.files.get("audio")

    if not audio_file:
        return jsonify({"error": "No audio file provided"}), 400

    duration, error = get_audio_duration_from_form_file(audio_file)

    if error:
        return jsonify({"error": error}), 400

    return jsonify({"message": "Audio processed successfully", "duration_minutes": duration}), 200


@app.route('/translate-with-google', methods=['POST'])
@require_api_key
def translate_text_endpoint():
    """Translates text to the target language"""
    try:
        logging.info("Received request to translate text with Google")

        data = request.get_json()
        if not data or "text" not in data or "target_language" not in data:
            return jsonify({'error': 'Missing text or target_language in request'}), 400
        
        text = data['text']
        target_language = data['target_language'] or "en"

        # Ensure text is a list
        if not isinstance(text, list):
            text = [text]

        translated_text_response = translate_text_google(text, target_language)
        translated_text = translated_text_response["joined_translated_text"]
        translated_text_list = translated_text_response["translated_text_list"]
        return jsonify({'translated_text': translated_text, 'translated_text_list': translated_text_list})

    except Exception as e:
        logging.error(f"Error in translate_text: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/translate-with-openai', methods=['POST'])
@require_api_key
def translate_text_with_openai_endpoint():
    """Translates text to the target language"""
    try:
        logging.info("Received request to translate text with OpenAI")
        text = request.form.get("text")
        is_dev = request.form.get("isDev")
        is_local = request.form.get("isLocal")
        if not text:
            return jsonify({'error': 'Missing text in request'}), 400
        source_language = request.form.get("source_language")
        if not source_language:
            return jsonify({'error': 'Missing source_language in request'}), 400
        target_language = request.form.get("target_language")
        if not target_language:
            return jsonify({'error': 'Missing target_language in request'}), 400
        file_name = request.form.get("fileName")
        if not file_name:
            return jsonify({'error': 'Missing file_name in request'}), 400
        duration = request.form.get("duration")
        if not duration:
            return jsonify({'error': 'Missing duration in request'}), 400
        drive_id = request.form.get("driveId")
        if not drive_id:
            return jsonify({'error': 'Missing driveId in request'}), 400
        group_id = request.form.get("groupId")
        if not group_id:
            return jsonify({'error': 'Missing groupId in request'}), 400
        file_id = request.form.get("fileId")
        if not file_id:
            return jsonify({'error': 'Missing fileId in request'}), 400
        folder_id = request.form.get("folderId")
        if not folder_id:
            return jsonify({'error': 'Missing folderId in request'}), 400
        project_name = request.form.get("projectName")
        if not project_name:
            return jsonify({'error': 'Missing projectName in request'}), 400
        
        glossary_file = request.files.get("glossaryFile")
        corrections_file = request.files.get("correctionsFile")
        glossary_map, corrections_map = parse_glossary_and_corrections(glossary_file, corrections_file)

        if glossary_file:
            translated_text = translate_text_with_openai(text, source_language, target_language, glossary_map, corrections_map)
        else:
            translated_text = translate_text_with_openai(text, source_language, target_language)
        #make an http request to  https://hook.eu2.make.com/xjxlm9ehhdn16mhtfnp77sxpgidvagqe with form-data body
        if is_local == "true":
            return jsonify({'translated_text': translated_text})
        else:
            if is_dev == "true":
                url = "https://hook.eu2.make.com/62p3xl6a7nnr14y89i6av1bxapyvxpxn"
            else:
                url = "https://hook.eu2.make.com/xjxlm9ehhdn16mhtfnp77sxpgidvagqe"
            logging.info(f"URL: {url}")
            response = requests.post(
                url,
                data={
                    "translation": translated_text,
                    "transcription": text,
                    "fileName": file_name,
                    "duration": duration,
                    "driveId": drive_id,
                    "groupId": group_id,
                    "folderId": folder_id,
                    "fileId": file_id,
                    "projectName": project_name
                }
            )
            if response.status_code != 200:
                return jsonify({'error': 'Failed to send request to Make'}), 500
            
            #return jsonify({'translated_text': translated_text})
            return jsonify({'message': 'Request sent to Make'}), 200
    except Exception as e:
        logging.error(f"Error in translate_text_with_openai: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/translate-with-deepseek', methods=['POST'])
@require_api_key
def translate_text_with_deepseek_endpoint():
    """Translates text to target language using DeepSeek"""
    text = request.form.get("text")
    is_dev = request.form.get("isDev")
    is_local = request.form.get("isLocal")
    if is_dev == "true":
        url = "https://hook.eu2.make.com/62p3xl6a7nnr14y89i6av1bxapyvxpxn"
    else:
        url = "https://hook.eu2.make.com/xjxlm9ehhdn16mhtfnp77sxpgidvagqe"
    logging.info("Received request to translate text with DeepSeek")
    if not text:
        return jsonify({'error': 'Missing text in request'}), 400
    source_language = request.form.get("source_language")
    if not source_language:
        return jsonify({'error': 'Missing source_language in request'}), 400
    target_language = request.form.get("target_language")
    if not target_language:
        return jsonify({'error': 'Missing target_language in request'}), 400
    file_name = request.form.get("fileName")
    if not file_name:
        return jsonify({'error': 'Missing file_name in request'}), 400
    duration = request.form.get("duration")
    if not duration:
        return jsonify({'error': 'Missing duration in request'}), 400
    drive_id = request.form.get("driveId")
    if not drive_id:
        return jsonify({'error': 'Missing driveId in request'}), 400
    group_id = request.form.get("groupId")
    if not group_id:
        return jsonify({'error': 'Missing groupId in request'}), 400
    file_id = request.form.get("fileId")
    if not file_id:
        return jsonify({'error': 'Missing fileId in request'}), 400
    folder_id = request.form.get("folderId")
    if not folder_id:
        return jsonify({'error': 'Missing folderId in request'}), 400
    project_name = request.form.get("projectName")
    if not project_name:
        return jsonify({'error': 'Missing projectName in request'}), 400
            
    glossary_file = request.files.get("glossaryFile")
    corrections_file = request.files.get("correctionsFile")
    glossary_map, corrections_map = parse_glossary_and_corrections(glossary_file, corrections_file)
    try:
        if glossary_file:
            translated_text = translate_text_with_deepseek(text, source_language, target_language, glossary_map, corrections_map)
        else:
            translated_text = translate_text_with_deepseek(text, source_language, target_language)
        #make an http request to  https://hook.eu2.make.com/xjxlm9ehhdn16mhtfnp77sxpgidvagqe with form-data body
        if is_local == "true":
            return jsonify({'translated_text': translated_text})
        else:
            logging.info(f"URL: {url}")
            response = requests.post(
                url,
                data={
                    "translation": translated_text,
                    "transcription": text,
                    "fileName": file_name,
                    "duration": duration,
                    "driveId": drive_id,
                    "groupId": group_id,
                    "folderId": folder_id,
                    "fileId": file_id,
                    "projectName": project_name
                }
            )
            if response.status_code != 200:
                return jsonify({'error': 'Failed to send request to Make'}), 500
            
            #return jsonify({'translated_text': translated_text})
            return jsonify({'message': 'Request sent to Make'}), 200
    except Exception as e:
        logging.error(f"Error in translate_text_with_openai: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route("/transcribe-google", methods=["POST"])
def transcribe_endpoint():
    """API endpoint to transcribe an audio file."""
    try:
        file = request.files['file']
        gcs_uri = request.form.get("gcs_uri")  # Optional GCS URI
        language_code = request.form.get("language", "en-US")
        temp_path = ''
        if not file and not gcs_uri:
            return jsonify({"error": "No file or GCS URI provided"}), 400

        if file:
            # Save the file temporarily
            temp_path = f"/tmp/{file.filename}"
            file.save(temp_path)

            # Upload to GCS
            bucket_name = "your-gcs-bucket"  # Replace with your actual bucket
            gcs_uri = upload_to_gcs(temp_path, bucket_name, file.filename)

        # Perform transcription
        transcription = transcribe_audio_openai(gcs_uri, language_code)
        return jsonify({"transcription": transcription})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    finally:
        # Clean up temporary files
        if os.path.exists(temp_path):
            os.remove(temp_path)
            #delete from bucket
            if gcs_uri:
                delete_from_gcs(gcs_uri)

           
@app.route("/transcript-with-whisper", methods=["POST"])
@require_api_key
def transcript_with_whisper_endpoint():
    """
    API endpoint to process and transcribe audio files using OpenAI Whisper.
    """
    logging.debug("Received request to process audio")
    audio_file = request.files.get("audio")
    language = request.form.get("language", "en")

    if not audio_file:
        return jsonify({"error": "No file uploaded"}), 400

    input_path = os.path.join("/tmp", audio_file.filename)
    wav_path = os.path.join("/tmp", f"{os.path.splitext(audio_file.filename)[0]}.wav")

    try:
        # Save the uploaded file
        logging.debug(f"Saving uploaded file to {input_path}")
        audio_file.save(input_path)

        # Convert to WAV if not already a WAV file
        if not input_path.lower().endswith(".wav"):
            logging.debug(f"Converting {input_path} to WAV format")
            convert_to_wav(input_path, wav_path)
        else:
            wav_path = input_path  # If already a WAV file, use as is
            logging.debug(f"Using existing WAV file: {wav_path}")
            
        # Process the WAV file using OpenAI Whisper API
        logging.debug(f"Transcripting WAV file whith whisper: {wav_path}")
        whisper_transcription = transcript_with_whisper_large_files(wav_path, "/tmp/temp", language)  # Ensure this uses the updated process_file

        # Check if transcription is valid
        if not whisper_transcription or not isinstance(whisper_transcription, str):
            logging.error("Transcription failed or returned invalid data")
            return jsonify({"error": "Transcription failed or returned invalid data"}), 500
        logging.debug(f"Whisper transcription: {whisper_transcription}")


        # Clean up uploaded and converted files
        if input_path != wav_path:  # Only delete input if converted
            logging.debug(f"Removing uploaded file: {input_path}")
            os.remove(input_path)
        logging.debug(f"Removing WAV file: {wav_path}")
        os.remove(wav_path)

        logging.info("Processing complete, returning transcription")


        result = {"message": "Processing complete", "transcription": whisper_transcription}
        return jsonify(result)

    except Exception as e:
        logging.error(f"Error processing audio: {e}")
        return jsonify({"error": str(e)}), 500

    finally:
        # Clean up temporary files
        temp_dir = "/tmp/temp"
        if os.path.exists(temp_dir):
            logging.debug(f"Cleaning up temporary files in {temp_dir}")
            for file in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, file))


@app.route("/text-to-file", methods=["POST"])
def text_to_file():
    """
    API endpoint to create a text file from input text.
    """
    logging.debug("Received request to create text file")
    
    try:
        # Get JSON data from request safely
        data = request.get_data(as_text=True)
        logging.debug(f"Raw request data: {data}")

        text = request.form.get("text") or request.json.get("text")
        fileName = request.form.get("fileName") or request.json.get("fileName")

        # Validate required fields
        if not text:
            logging.error("No text provided")
            return jsonify({"error": "No text provided"}), 400
        
        if not fileName:
            logging.error("No fileName provided")
            return jsonify({"error": "No fileName provided"}), 400

        filename = f"{fileName}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

        # Create file-like object in memory
        file_obj = io.BytesIO()
        file_obj.write(text.encode('utf-8'))  # Ensure proper UTF-8 encoding
        file_obj.seek(0)

        # Return the file
        return send_file(
            file_obj,
            mimetype='text/plain',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error(f"Error creating text file: {e}")
        return jsonify({"error": str(e)}), 500
    
@app.route('/sentiment-analysis', methods=['POST'])
@require_api_key
def analyze_sentiment_and_queries():
    try:
        excel_file = request.files.get('file')
        text_to_analyze = request.form.get('text')
        
        if not excel_file:
            return jsonify({"error": "No Excel file provided."}), 400
        if not text_to_analyze:
            return jsonify({"error": "No text provided for analysis."}), 400

        # Process queries (Sheet 1)
        df_queries = load_excel_file(excel_file)
        responses = process_queries(df_queries, text_to_analyze)
        df_queries['Response'] = responses

        # Run sentiment analysis using the Hospital_Reviews model
        sentiment_results = run_sentiment_analysis(text_to_analyze)
        df_sentiment_details = create_sentiment_details_df(sentiment_results)
        df_sentiment_summary = create_sentiment_summary_df(df_sentiment_details)

        # Generate the multi-sheet Excel file
        output = generate_multi_sheet_excel(df_queries, df_sentiment_details, df_sentiment_summary)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name="analyzed_data.xlsx"
        )
    except Exception as e:
        logging.error(f"Error in analyze_sentiment_and_queries endpoint: {e}")
        return jsonify({"error": str(e)}), 500
    
@app.route('/generate-excel', methods=['POST'])
@require_api_key
def generate_excel():
    try:
        # Parse the incoming JSON
        data = request.json
        if not data or not is_valid_json(request.data.decode('utf-8')):
            return jsonify({"error": "Invalid JSON format 1"}), 400
        if "sheets" not in data:
            return jsonify({"error": "Invalid JSON format 2"}), 400

        logging.info(f"DATA: {data}")

        # Create a new Excel workbook
        workbook = Workbook()
        # Remove default sheet
        workbook.remove(workbook.active)

        # Process each sheet in the JSON
        for sheet_data in data["sheets"]:
            sheet_name = sheet_data.get("name", "Sheet")
            sheet_data_rows = sheet_data.get("data", [])

            # Create a new sheet
            sheet = workbook.create_sheet(title=sheet_name)

            # Write the rows directly to the sheet
            for row in sheet_data_rows:
                sheet.append(row)

        # Save workbook to a BytesIO stream
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Return the Excel file
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name="data_analysis.xlsx")

    except Exception as e:
        # Log the full exception stack trace for debugging
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/generate-word", methods=["POST"])
@require_api_key
def generate_word():
    """
    API endpoint to generate a Word document from FormData (text + filename).
    """
    try:
        content = request.form.get("text", "")
        filename = request.form.get("fileName", "file.docx")

        file_path = create_word_document(content, filename)
        return send_file(file_path, as_attachment=True)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/log-audio-usage", methods=["POST"])
@require_api_key
def log_audio_usage():
    """
    Endpoint per elaborare l'audio e registrare la durata.
    """
    try:
        user_code = request.form.get("user_code")  # Codice utente per identificare il file
        filename = request.form.get("fileName")
        duration = request.form.get("duration")

        if not user_code or not filename or not duration:
            return jsonify({"error": "Manca il codice utente, il nome del file o la durata"}), 400

        # Log the data
        result = log_audio_processing(user_code, filename, duration)

        if result is None:
            return jsonify({"error": "Error during Google Sheets logging"}), 500

        return jsonify({
            "message": "File elaborato con successo",
            "user_code": user_code,
            "filename": filename,
            "duration": duration,
            "cost_per_minute": f"{RATE_PER_MINUTE:.2f}",
            "total_cost": f"{(float(duration) * RATE_PER_MINUTE):.2f}",
            "Billed": 'YES'
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/transcribe-with-assemblyai", methods=["POST"])
@require_api_key
def transcribe_audio_with_assemblyai():
    if "audio" not in request.files:
        return jsonify({"error": "No audio file provided"}), 400
    audio_file = request.files.get("audio")
    is_dev = request.form.get("isDev")
    is_local = request.form.get("isLocal")
    source_language = request.form.get("sourceLanguage")
    if not source_language:
        return jsonify({'error': 'Missing source_language in request'}), 400
    target_language = request.form.get("targetLanguage")
    if not target_language:
        return jsonify({'error': 'Missing target_language in request'}), 400
    file_name = request.form.get("fileName")
    if not file_name:
        return jsonify({'error': 'Missing file_name in request'}), 400
    duration = request.form.get("duration")
    if not duration:
        return jsonify({'error': 'Missing duration in request'}), 400
    drive_id = request.form.get("driveId")
    if not drive_id:
        return jsonify({'error': 'Missing driveId in request'}), 400
    group_id = request.form.get("groupId")
    if not group_id:
        return jsonify({'error': 'Missing groupId in request'}), 400
    file_id = request.form.get("fileId")
    if not file_id:
        return jsonify({'error': 'Missing fileId in request'}), 400
    folder_id = request.form.get("folderId")
    if not folder_id:
        return jsonify({'error': 'Missing folderId in request'}), 400
    project_name = request.form.get("projectName")
    if not project_name:
        return jsonify({'error': 'Missing projectName in request'}), 400
    
    temp_path = f"temp_{audio_file.filename}"
    audio_file.save(temp_path)
    logging.info(f"SOURCE LANGUAGE: {source_language}")
    #if source_language is ina list use a model otherwise another one
    if source_language in ["en", "en_au", "en_uk", "en_us", "es", "fr", "de", "it", "pt", "nl", "hi", "ja", "zh", "fi", "ko", "pl", "ru", "tr", "uk", "vi"]:
        logging.info("Using best model")
        model = aai.SpeechModel.best
        speaker_labels = True
    else:
        logging.info("Using nano model")
        model = aai.SpeechModel.nano
        speaker_labels = False
    try:
        logging.info(f"SPEAKER LABELS: {speaker_labels}")
        logging.info(f"MODEL: {model}")
        logging.info(f"SOURCE LANGUAGE: {source_language}")
        logging.info(f"TARGET LANGUAGE: {target_language}")
        logging.info(f"FILE NAME: {file_name}")
        logging.info(f"DURATION: {duration}")
        logging.info(f"DRIVE ID: {drive_id}")
        logging.info(f"GROUP ID: {group_id}")
        logging.info(f"FILE ID: {file_id}")
        logging.info(f"FOLDER ID: {folder_id}")
        # Configure for Chinese with Nano model and speaker diarization
        config = aai.TranscriptionConfig(
            language_code=source_language,
            speech_model=model,
            speaker_labels=speaker_labels
        )
        
        transcriber = aai.Transcriber()
        transcript = transcriber.transcribe(temp_path, config)
        if transcript.status == aai.TranscriptStatus.error:
            return jsonify({"error": transcript.error}), 500
        
        # Format the transcript with speaker labels as requested
        if speaker_labels:
            formatted_transcript = ""
            for utterance in transcript.utterances:
                formatted_transcript += f"speaker{utterance.speaker}: {utterance.text}\n"
        else:
            formatted_transcript = transcript.text
        #return jsonify({
        #    "transcription": formatted_transcript,
        #})
        logging.info(f"FORMATTED TRANSCRIPT: {formatted_transcript}")
        if is_local == "true":
            return jsonify({'transcription': formatted_transcript})
        if is_dev == "true":
            url = "https://hook.eu2.make.com/1qn49rif17gctwp53zee3xbjb6aqvbko"
        else:
            url = "https://hook.eu2.make.com/qcc3jfwa2stoz8xqzjvap6581hqyl2oy"
        logging.info(f"URL: {url}")
        response = requests.post(
            url,
            data={
                "transcription": formatted_transcript,
                "fileName": file_name,
                "duration": duration,
                "driveId": drive_id,
                "groupId": group_id,
                "folderId": folder_id,
                "fileId": file_id,
                "projectName": project_name,
                "sourceLanguage": source_language
            }
        )
        if response.status_code != 200:
            return jsonify({'error': 'Failed to send request to Make'}), 500
        return jsonify({'message': 'Request sent to Make'}), 200
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@app.route('/generate-monthly-report', methods=['POST'])
@require_api_key
def generate_monthly_report():
    try:
        # Fetch usage data
        user_code = request.form.get("user_code")
        data = get_usage_data(user_code)

        # Filter data where 'Billed' is 'YES'
        billed_data = [record for record in data if record.get('Billed') == 'YES']

        # Filter data for the current month
        current_month = datetime.datetime.now().month
        monthly_data = [
            record for record in billed_data
            if datetime.datetime.strptime(record['Data e ora'], '%Y-%m-%d %H:%M:%S').month == current_month
        ]

        # Generate PDF
        pdf_gen = PDFGenerator('Monthly Usage Report')
        pdf_gen.add_table(monthly_data)
        pdf_filename = f'monthly_report_{current_month}.pdf'
        pdf_gen.save_pdf(pdf_filename)

        # Schedule file deletion after sending the response
        @after_this_request
        def remove_file(response):
            try:
                os.remove(pdf_filename)
                print(f"Deleted file: {pdf_filename}")  # Debugging
            except Exception as e:
                print(f"Error deleting file: {str(e)}")
            return response

        return send_file(pdf_filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate-billing-document', methods=['POST'])
@require_api_key
def generate_billing_document():
    try:
        # Fetch usage data
        user_code = request.form.get("user_code")
        data = get_usage_data(user_code)

        # Filter data for the current month
        current_month = datetime.datetime.now().month
        monthly_data = [record for record in data if datetime.datetime.strptime(record['Data e ora'], '%Y-%m-%d %H:%M:%S').month == current_month]

        # Calculate total cost
        total_cost = sum(float(record['Costo Totale (€)'].replace(',', '.')) for record in monthly_data)

        # Generate PDF Invoice
        pdf_gen = PDFGenerator('Monthly Billing Document')
        pdf_gen.add_table(monthly_data)
        pdf_gen.pdf.ln(10)
        pdf_gen.pdf.cell(0, 10, f'Total Cost: €{total_cost:.2f}', 0, 1, 'R')
        pdf_filename = f'billing_document_{current_month}.pdf'
        pdf_gen.save_pdf(pdf_filename)

        return send_file(pdf_filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)